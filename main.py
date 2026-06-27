import argparse
import csv
import os
import re
import sys
from dotenv import load_dotenv

load_dotenv()

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import pickle

SPREADSHEET_ID = os.getenv("SPREADSHEET_ID", "")
SHEET_NAME = os.getenv("SHEET_NAME", "Sheet1")
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
TOKEN_PATH = os.path.join(os.path.dirname(__file__), "token.pickle")
CREDENTIALS_PATH = os.path.join(os.path.dirname(__file__), "credentials.json")

# Exact handles (without @, case-insensitive)
TG_BLACKLIST_EXACT: set[str] = {
    "sem0nch1k",
    "a_rtem_sch",
}

# Prefixes (handle starts with this string)
TG_BLACKLIST_PREFIXES: list[str] = [
    "coranmiel",
]

# Regex patterns (matched against full handle)
TG_BLACKLIST_PATTERNS: list[str] = [
    # r"test\d+",
]

STATUS_MAP = {
    "Оплачено": ("Оплатил", "008000"),
    "Ожидает оплаты": ("Зарегистрировался", "ADD8E6"),
    "Отменено": ("Напомнили", "8B0000"),
    "Просрочено": ("Напомнили", "8B0000"),
    "Изменено": ("Отдали промокод", "800080"),
}

SORT_ORDER = {
    "Оплачено": 0,
    "Ожидает оплаты": 1,
    "Изменено": 2,
    "Отменено": 3,
    "Просрочено": 4,
}

OUTPUT_HEADERS = [
    "Статус регистрации",
    "Сайт актуален?",
    "ID регистрации",
    "ФИО",
    "Email",
    "Telegram",
    "Волна",
    "Билет",
    "Статус",
    "Тип оплаты",
    "Полная сумма",
    "Скидка",
    "Итого",
    "Время регистрации",
    "Время подтверждения",
    "Получатель (имя)",
    "Получатель (банк)",
    "Реквизиты перевода",
]

CSV_TO_INTERNAL = {
    "ID регистрации": "ID регистрации",
    "ФИО": "ФИО",
    "Email": "Email",
    "Telegram": "Telegram",
    "Волна": "Волна",
    "Билет": "Билет",
    "Статус": "Статус",
    "Тип оплаты": "Тип оплаты",
    "Полная сумма": "Полная сумма",
    "Скидка": "Скидка",
    "Итого": "Итого",
    "Время регистрации": "Время регистрации",
    "Время подтверждения": "Время подтверждения",
    "Получатель (имя)": "Получатель (имя)",
    "Получатель (банк)": "Получатель (банк)",
    "Реквизиты перевода": "Реквизиты перевода",
}


def parse_csv(path: str) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        return list(reader)


def normalize_tg(handle: str) -> str:
    return handle.lstrip("@").lower().strip()


def deduplicate(rows: list[dict]) -> list[dict]:
    """Keep first entry per ФИО (CSV is ordered newest→oldest, so first = most recent)."""
    seen: dict[str, dict] = {}
    for row in rows:
        name = row.get("ФИО", "").strip()
        if name and name not in seen:
            seen[name] = row
    return list(seen.values())


def is_blacklisted(handle: str) -> bool:
    tg = normalize_tg(handle)
    if tg in TG_BLACKLIST_EXACT:
        return True
    if any(tg.startswith(p.lower()) for p in TG_BLACKLIST_PREFIXES):
        return True
    if any(re.fullmatch(p, tg) for p in TG_BLACKLIST_PATTERNS):
        return True
    return False


def filter_blacklist(rows: list[dict]) -> list[dict]:
    return [row for row in rows if not is_blacklisted(row.get("Telegram", ""))]


def _to_int(val: str):
    try:
        return int(val.strip())
    except (ValueError, AttributeError):
        return val


def build_output_rows(rows: list[dict]) -> list[dict]:
    out = []
    for row in rows:
        status_raw = row.get("Статус", "").strip()
        label, _ = STATUS_MAP.get(status_raw, (status_raw, "FFFFFF"))
        record = {
            "Статус регистрации": label,
            "Сайт актуален?": True,
            "ID регистрации": row.get("ID регистрации", ""),
            "ФИО": row.get("ФИО", ""),
            "Email": row.get("Email", ""),
            "Telegram": row.get("Telegram", ""),
            "Волна": row.get("Волна", ""),
            "Билет": row.get("Билет", ""),
            "Статус": status_raw,
            "Тип оплаты": row.get("Тип оплаты", ""),
            "Полная сумма": _to_int(row.get("Полная сумма", "")),
            "Скидка": _to_int(row.get("Скидка", "")),
            "Итого": _to_int(row.get("Итого", "")),
            "Время регистрации": row.get("Время регистрации", ""),
            "Время подтверждения": row.get("Время подтверждения", ""),
            "Получатель (имя)": row.get("Получатель (имя)", ""),
            "Получатель (банк)": row.get("Получатель (банк)", ""),
            "Реквизиты перевода": row.get("Реквизиты перевода", ""),
        }
        out.append((record, status_raw))
    return out


def write_xlsx(output_rows: list[tuple], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Регистрации"

    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_font = Font(bold=True)

    for col, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    status_fill = {
        "Оплачено": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "Ожидает оплаты": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "Отменено": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "Просрочено": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        "Изменено": PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),
    }

    for row_idx, (record, status_raw) in enumerate(output_rows, start=2):
        for col, header in enumerate(OUTPUT_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col, value=record.get(header, ""))
            if header == "Статус регистрации" and status_raw in status_fill:
                cell.fill = status_fill[status_raw]

    wb.save(path)


def get_google_creds() -> Credentials:
    creds = None
    if os.path.exists(TOKEN_PATH):
        with open(TOKEN_PATH, "rb") as f:
            creds = pickle.load(f)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_PATH):
                print(
                    f"Файл credentials.json не найден по пути {CREDENTIALS_PATH}.\n"
                    "Скачай его из Google Cloud Console (OAuth 2.0 Desktop) и положи рядом со скриптом.",
                    file=sys.stderr,
                )
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_PATH, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_PATH, "wb") as f:
            pickle.dump(creds, f)
    return creds


def push_to_sheets(output_rows: list[tuple]):
    creds = get_google_creds()
    service = build("sheets", "v4", credentials=creds)
    sheet = service.spreadsheets()

    # Clear the sheet first
    sheet.values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=SHEET_NAME,
    ).execute()

    values = [OUTPUT_HEADERS]
    for record, _ in output_rows:
        row = []
        for h in OUTPUT_HEADERS:
            v = record.get(h, "")
            row.append("TRUE" if v is True else ("FALSE" if v is False else v))
        values.append(row)

    sheet.values().update(
        spreadsheetId=SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": values},
    ).execute()

    # Apply background colors via batchUpdate
    requests = []

    # Find sheet ID
    meta = sheet.get(spreadsheetId=SPREADSHEET_ID).execute()
    sheet_id = None
    for s in meta["sheets"]:
        if s["properties"]["title"] == SHEET_NAME:
            sheet_id = s["properties"]["sheetId"]
            break

    if sheet_id is None:
        print(f"Лист '{SHEET_NAME}' не найден в таблице.", file=sys.stderr)
        return


    # Checkbox for column B (Сайт актуален?)
    requests.append({
        "setDataValidation": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 1,
                "endRowIndex": len(output_rows) + 1,
                "startColumnIndex": 1,
                "endColumnIndex": 2,
            },
            "rule": {
                "condition": {"type": "BOOLEAN"},
                "showCustomUi": True,
            },
        }
    })

    if requests:
        sheet.batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": requests},
        ).execute()

    print(f"Данные ({len(output_rows)} строк) загружены в Google Sheets: '{SHEET_NAME}'")


def main():
    parser = argparse.ArgumentParser(description="Обработка регистраций из CSV")
    parser.add_argument("input", help="Путь к входному CSV файлу")
    parser.add_argument("output", help="Путь к выходному XLSX файлу")
    parser.add_argument(
        "--no-sheets",
        action="store_true",
        help="Не загружать данные в Google Sheets",
    )
    args = parser.parse_args()

    rows = parse_csv(args.input)
    rows = filter_blacklist(rows)
    rows = deduplicate(rows)
    output_rows = build_output_rows(rows)
    output_rows.sort(key=lambda x: SORT_ORDER.get(x[1], 99))

    write_xlsx(output_rows, args.output)
    print(f"XLSX сохранён: {args.output} ({len(output_rows)} строк)")

    if not args.no_sheets:
        push_to_sheets(output_rows)


if __name__ == "__main__":
    main()
